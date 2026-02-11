from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.http import HttpResponse
from datetime import datetime, date, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import AttendanceGroup, UserGroup, AttendanceLog
from users.models import User


def user_is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.is_admin()


@login_required
def employee_dashboard(request):
    """
    Employee dashboard - Simple clock in/out interface
    """
    if request.user.is_admin():
        return redirect('admin_dashboard')
    
    today = timezone.localdate()
    user = request.user
    
    # Get user's attendance groups
    user_groups = UserGroup.objects.filter(user=user).select_related('group')
    
    # Check if today is an allowed day for any of the user's groups
    is_allowed_today = False
    allowed_groups = []
    
    for ug in user_groups:
        if ug.group.is_day_allowed(today):
            is_allowed_today = True
            allowed_groups.append(ug.group.name)
    
    # Get today's attendance log if exists
    today_log = AttendanceLog.objects.filter(user=user, date=today).first()
    
    # Determine button state
    can_check_in = is_allowed_today and (today_log is None or today_log.check_in is None)
    can_check_out = today_log is not None and today_log.check_in is not None and today_log.check_out is None
    is_finished = today_log is not None and today_log.is_complete()
    
    # Get recent attendance history (last 7 days)
    recent_logs = AttendanceLog.objects.filter(
        user=user,
        date__gte=today - timedelta(days=7)
    ).order_by('-date')
    
    context = {
        'today': today,
        'current_time': timezone.localtime(),
        'is_allowed_today': is_allowed_today,
        'allowed_groups': allowed_groups,
        'today_log': today_log,
        'can_check_in': can_check_in,
        'can_check_out': can_check_out,
        'is_finished': is_finished,
        'recent_logs': recent_logs,
        'user_groups': user_groups,
    }
    
    return render(request, 'attendance/employee_dashboard.html', context)


@login_required
def clock_in(request):
    """Handle clock in action"""
    if request.method != 'POST':
        return redirect('employee_dashboard')
    
    user = request.user
    today = timezone.localdate()
    now = timezone.localtime().time()
    
    # Verify user has permission to clock in today
    user_groups = UserGroup.objects.filter(user=user).select_related('group')
    is_allowed = any(ug.group.is_day_allowed(today) for ug in user_groups)
    
    if not is_allowed:
        messages.error(request, f'You are not allowed to clock in on {today.strftime("%A")}.')
        return redirect('employee_dashboard')
    
    # Check if already clocked in today
    log, created = AttendanceLog.objects.get_or_create(
        user=user,
        date=today,
        defaults={'check_in': now}
    )
    
    if not created:
        if log.check_in is not None:
            messages.warning(request, 'You have already clocked in today.')
        else:
            log.check_in = now
            log.save()
            messages.success(request, f'Successfully clocked in at {now.strftime("%I:%M %p")}.')
    else:
        messages.success(request, f'Successfully clocked in at {now.strftime("%I:%M %p")}.')
    
    return redirect('employee_dashboard')


@login_required
def clock_out(request):
    """Handle clock out action"""
    if request.method != 'POST':
        return redirect('employee_dashboard')
    
    user = request.user
    today = timezone.localdate()
    now = timezone.localtime().time()
    
    try:
        log = AttendanceLog.objects.get(user=user, date=today)
        
        if log.check_in is None:
            messages.error(request, 'You must clock in before clocking out.')
        elif log.check_out is not None:
            messages.warning(request, 'You have already clocked out today.')
        else:
            log.check_out = now
            log.save()
            
            total_hours = log.get_total_hours()
            messages.success(
                request,
                f'Successfully clocked out at {now.strftime("%I:%M %p")}. '
                f'Total hours: {total_hours:.2f}'
            )
    
    except AttendanceLog.DoesNotExist:
        messages.error(request, 'No clock in record found for today.')
    
    return redirect('employee_dashboard')


@login_required
def admin_dashboard(request):
    """
    Admin dashboard - Overview and management
    """
    if not user_is_admin(request.user):
        return redirect('employee_dashboard')
    
    today = timezone.localdate()
    
    # Get today's active employees (clocked in, not yet out)
    active_today = AttendanceLog.objects.filter(
        date=today,
        check_in__isnull=False,
        check_out__isnull=True
    ).select_related('user').order_by('check_in')
    
    # Get completed logs for today
    completed_today = AttendanceLog.objects.filter(
        date=today,
        check_in__isnull=False,
        check_out__isnull=False
    ).select_related('user').order_by('-check_out')
    
    # Statistics
    total_users = User.objects.filter(is_active=True, role='EMPLOYEE').count()
    total_groups = AttendanceGroup.objects.count()
    clocked_in_count = active_today.count()
    completed_count = completed_today.count()
    
    context = {
        'today': today,
        'active_today': active_today,
        'completed_today': completed_today,
        'total_users': total_users,
        'total_groups': total_groups,
        'clocked_in_count': clocked_in_count,
        'completed_count': completed_count,
    }
    
    return render(request, 'attendance/admin_dashboard.html', context)


@login_required
def manage_groups(request):
    """Manage attendance groups"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    groups = AttendanceGroup.objects.all().annotate(
        user_count=Count('group_users')
    )
    
    context = {'groups': groups}
    return render(request, 'attendance/manage_groups.html', context)


@login_required
def create_group(request):
    """Create new attendance group"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        selected_days = request.POST.getlist('days')
        
        if not name or not selected_days:
            messages.error(request, 'Please provide a name and select at least one day.')
            return redirect('create_group')
        
        # Convert day strings to integers
        allowed_days = [int(day) for day in selected_days]
        
        try:
            group = AttendanceGroup.objects.create(
                name=name,
                allowed_days=allowed_days
            )
            messages.success(request, f'Group "{group.name}" created successfully.')
            return redirect('manage_groups')
        except Exception as e:
            messages.error(request, f'Error creating group: {str(e)}')
    
    weekdays = AttendanceGroup.WEEKDAYS
    context = {'weekdays': weekdays}
    return render(request, 'attendance/create_group.html', context)


@login_required
def edit_group(request, group_id):
    """Edit attendance group"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    group = get_object_or_404(AttendanceGroup, id=group_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        selected_days = request.POST.getlist('days')
        
        if not name or not selected_days:
            messages.error(request, 'Please provide a name and select at least one day.')
        else:
            group.name = name
            group.allowed_days = [int(day) for day in selected_days]
            group.save()
            messages.success(request, f'Group "{group.name}" updated successfully.')
            return redirect('manage_groups')
    
    weekdays = AttendanceGroup.WEEKDAYS
    context = {
        'group': group,
        'weekdays': weekdays,
    }
    return render(request, 'attendance/edit_group.html', context)


@login_required
def delete_group(request, group_id):
    """Delete attendance group"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    group = get_object_or_404(AttendanceGroup, id=group_id)
    
    if request.method == 'POST':
        group_name = group.name
        group.delete()
        messages.success(request, f'Group "{group_name}" deleted successfully.')
        return redirect('manage_groups')
    
    context = {'group': group}
    return render(request, 'attendance/delete_group.html', context)


@login_required
def manage_users(request):
    """Manage users and their group assignments"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    users = User.objects.filter(is_active=True).prefetch_related('user_groups__group')
    
    context = {'users': users}
    return render(request, 'attendance/manage_users.html', context)


@login_required
def assign_user_to_group(request, user_id):
    """Assign user to attendance groups"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        selected_groups = request.POST.getlist('groups')
        
        # Remove all current assignments
        UserGroup.objects.filter(user=user).delete()
        
        # Add new assignments
        for group_id in selected_groups:
            group = AttendanceGroup.objects.get(id=group_id)
            UserGroup.objects.create(user=user, group=group)
        
        messages.success(request, f'Groups updated for {user.get_full_name() or user.username}.')
        return redirect('manage_users')
    
    all_groups = AttendanceGroup.objects.all()
    user_group_ids = set(user.user_groups.values_list('group_id', flat=True))
    
    context = {
        'user': user,
        'all_groups': all_groups,
        'user_group_ids': user_group_ids,
    }
    return render(request, 'attendance/assign_user_groups.html', context)


@login_required
def reports(request):
    """View and export attendance reports"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    user_id = request.GET.get('user')
    
    # Build query
    logs = AttendanceLog.objects.select_related('user').order_by('-date', 'user__username')
    
    if start_date:
        logs = logs.filter(date__gte=start_date)
    if end_date:
        logs = logs.filter(date__lte=end_date)
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    # Get all users for filter dropdown
    all_users = User.objects.filter(is_active=True, role='EMPLOYEE').order_by('username')
    
    context = {
        'logs': logs,
        'all_users': all_users,
        'start_date': start_date,
        'end_date': end_date,
        'selected_user': user_id,
    }
    return render(request, 'attendance/reports.html', context)


@login_required
def export_excel(request):
    """Export attendance records to Excel"""
    if not user_is_admin(request.user):
        messages.error(request, 'Access denied.')
        return redirect('employee_dashboard')
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    user_id = request.GET.get('user')
    
    # Build query
    logs = AttendanceLog.objects.select_related('user').order_by('date', 'user__username')
    
    if start_date:
        logs = logs.filter(date__gte=start_date)
    if end_date:
        logs = logs.filter(date__lte=end_date)
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    headers = ['User', 'Date', 'Check In', 'Check Out', 'Total Hours']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for log in logs:
        ws.append([
            log.user.get_full_name() or log.user.username,
            log.date.strftime('%Y-%m-%d'),
            log.check_in.strftime('%I:%M %p') if log.check_in else 'N/A',
            log.check_out.strftime('%I:%M %p') if log.check_out else 'N/A',
            f"{log.get_total_hours():.2f}" if log.is_complete() else 'N/A',
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"attendance_report_{timezone.localdate().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
